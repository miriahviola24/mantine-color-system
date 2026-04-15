#!/usr/bin/env python3
"""
Build color-system-mantine-expansion.xlsx + update JSON files with magenta.
"""
import json, os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = Path("/Users/mviola/Library/Application Support/Claude/local-agent-mode-sessions/ac4f0e3f-0aa8-4bd1-a6ef-42a6c050736e/7efedaae-5a9f-43a8-ae3e-8e33a16adf98/local_75ac7152-fd82-44e0-9b8b-351f5ed40111/outputs")

# ───────────────────────────────────────────────
# Magenta scale computation
# ───────────────────────────────────────────────
def h2r(h):
    h = h.lstrip('#')
    return (int(h[0:2],16), int(h[2:4],16), int(h[4:6],16))

def r2h(r,g,b):
    return '#{:02X}{:02X}{:02X}'.format(round(r), round(g), round(b))

def pwlerp(anchors, t):
    anchors = sorted(anchors, key=lambda x: x[0])
    if t <= anchors[0][0]: return anchors[0][1]
    if t >= anchors[-1][0]: return anchors[-1][1]
    for i in range(len(anchors)-1):
        t0,c0 = anchors[i]; t1,c1 = anchors[i+1]
        if t0 <= t <= t1:
            u = (t-t0)/(t1-t0)
            return tuple(c0[j]+(c1[j]-c0[j])*u for j in range(3))
    return anchors[-1][1]

L_ANCH = [(0,h2r("#FFF1FC")), (2/9,h2r("#F9B0EE")), (5/9,h2r("#AC037A")),
           (7/9,h2r("#6A014A")), (1.0,h2r("#210018"))]
D_ANCH = [(0,h2r("#210018")), (2/9,h2r("#5C0047")), (5/9,h2r("#AC037A")),
           (7/9,h2r("#E276CE")), (1.0,h2r("#FFF1FC"))]

DESC = [
    "Lightest magenta tint \u2014 subtle background",
    "Very light \u2014 hover/selection bg",
    "Soft \u2014 secondary surface",
    "Border / divider",
    "Inactive border / chip",
    "Primary color (exact 500 value)",
    "Strong / pressed state",
    "Deep shade / filled badge",
    "Near-dark \u2014 high emphasis",
    "Darkest \u2014 near-black equivalent",
]

mag_l, mag_d = {}, {}
for i in range(10):
    t = i/9
    mag_l[str(i)] = {"value": r2h(*pwlerp(L_ANCH,t)), "type":"color", "description": DESC[i]}
    mag_d[str(i)] = {"value": r2h(*pwlerp(D_ANCH,t)), "type":"color", "description": DESC[i]+" (dark)"}

print("Magenta scale:")
for i in range(10):
    print(f"  {i}: L={mag_l[str(i)]['value']}  D={mag_d[str(i)]['value']}")

# ───────────────────────────────────────────────
# Update tokens-light.json
# ───────────────────────────────────────────────
with open(OUT/"tokens-light.json") as f: ld = json.load(f)
ld["magenta"] = mag_l
with open(OUT/"tokens-light.json","w") as f: json.dump(ld, f, indent=2)
print("\n✓ tokens-light.json updated")

# ───────────────────────────────────────────────
# Update tokens-dark.json
# ───────────────────────────────────────────────
with open(OUT/"tokens-dark.json") as f: dd = json.load(f)
dd["magenta"] = mag_d
with open(OUT/"tokens-dark.json","w") as f: json.dump(dd, f, indent=2)
print("✓ tokens-dark.json updated")

# ───────────────────────────────────────────────
# Update variables-import-export.json
# (apply 16 corrections + add magenta)
# ───────────────────────────────────────────────
CL = {"blue/2":"#82B8EB","blue/7":"#02396C","gray/1":"#E3E3E3","gray/2":"#C7C7C7",
      "gray/7":"#3A3A3A","gray/8":"#2D2D2D","orange/2":"#F2C7AF","orange/7":"#72482F",
      "purple/2":"#D2B1F7","purple/7":"#523178","green/2":"#A0CEB2","green/7":"#214F32",
      "yellow/2":"#EDDDAD","yellow/7":"#6E5E2D","red/2":"#EC958E","red/7":"#99251D"}
CD = {"blue/2":"#02396C","blue/7":"#82B8EB","gray/1":"#2D2D2D","gray/2":"#515151",
      "gray/7":"#C7C7C7","gray/8":"#E3E3E3","orange/2":"#72482F","orange/7":"#F2C7AF",
      "purple/2":"#523178","purple/7":"#D2B1F7","green/2":"#214F32","green/7":"#A0CEB2",
      "yellow/2":"#6E5E2D","yellow/7":"#EDDDAD","red/2":"#99251D","red/7":"#EC958E"}

with open(OUT/"variables-import-export.json") as f: vd = json.load(f)
for var in vd["collections"][0]["variables"]:
    n = var["name"]
    if n in CL: var["Light"] = CL[n]
    if n in CD: var["Dark"]  = CD[n]
for i in range(10):
    vd["collections"][0]["variables"].append({
        "name": f"magenta/{i}", "type":"COLOR", "description": DESC[i],
        "Light": mag_l[str(i)]["value"], "Dark": mag_d[str(i)]["value"]
    })
with open(OUT/"variables-import-export.json","w") as f: json.dump(vd, f, indent=2)
print("✓ variables-import-export.json updated (16 corrections + magenta)")

# ───────────────────────────────────────────────
# Group metadata
# ───────────────────────────────────────────────
GROUPS = [
    ("blue","Blue"), ("gray","Gray"), ("orange","Orange"), ("purple","Purple"),
    ("green","Green"), ("yellow","Yellow"), ("red","Red"),
    ("dataBlue","Data Blue"), ("dataTeal","Data Teal"), ("dataIndigo","Data Indigo"),
    ("dataAmber","Data Amber"), ("dataMint","Data Mint"), ("magenta","Magenta"),
]

OLD_NAMES = {
    "blue/0":"Blue 100","blue/2":"Blue 250","blue/5":"Brand / PrimaryBlue",
    "blue/7":"Blue 750","blue/9":"Blue 1000",
    "gray/0":"Neutral 100","gray/1":"Neutral 150","gray/2":"Neutral 250",
    "gray/5":"Neutral 500","gray/7":"Neutral 750","gray/8":"Neutral 850","gray/9":"Neutral 1000",
    "orange/0":"Orange 100","orange/2":"Orange 250","orange/5":"Orange 500",
    "orange/7":"Orange 750","orange/9":"Orange 1000",
    "purple/0":"Purple 100","purple/2":"Purple 250","purple/5":"Purple 500",
    "purple/7":"Purple 750","purple/9":"Purple 1000",
    "green/0":"Green 100","green/2":"Green 250","green/5":"Green 500",
    "green/7":"Green 750","green/9":"Green 1000",
    "yellow/0":"Yellow 100","yellow/2":"Yellow 250","yellow/5":"Yellow 500",
    "yellow/7":"Yellow 750","yellow/9":"Yellow 1000",
    "red/0":"Red 100","red/2":"Red 250","red/5":"Red 500",
    "red/7":"Red 750","red/9":"Red 1000",
    "dataAmber/5":"Targets 500","dataBlue/5":"Primary Data 500",
    "dataIndigo/5":"Projections 500","dataMint/5":"Upsides 500","dataTeal/5":"Secondary Data 500",
}

def old_name(key, step):
    token = f"{key}/{step}"
    if key == "magenta":
        # Mark pinned/anchored steps as NEW
        return "NEW" if step in (0,2,5,7,9) else "\u2014"
    return OLD_NAMES.get(token, "\u2014")

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color.lstrip('#'))

def accent(key):
    """Step-5 light color for a group."""
    return ld[key]["5"]["value"]

# ───────────────────────────────────────────────
# Build xlsx
# ───────────────────────────────────────────────
wb = openpyxl.Workbook()

# ─── Sheet 1: All Colors ───────────────────────
ws1 = wb.active
ws1.title = "All Colors"
ws1.freeze_panes = "A2"

HDR = ["Group","Token","Step","Old Name","Light Hex","Light Swatch","Dark Hex","Dark Swatch","Pinned?"]
ws1.append(HDR)
for col, w in enumerate([14,14,6,22,12,10,12,10,14], start=1):
    ws1.column_dimensions[get_column_letter(col)].width = w
ws1.row_dimensions[1].height = 22

hdr_font  = Font(name="Arial", bold=True, size=10)
body_font = Font(name="Arial", size=10)
hdr_fill  = fill("2D3748")
hdr_white = Font(name="Arial", bold=True, size=10, color="FFFFFF")
step5_fill = fill("FFFBE6")
pin_fill   = fill("FFE066")

for c in range(1, len(HDR)+1):
    cell = ws1.cell(1, c)
    cell.font  = hdr_white
    cell.fill  = hdr_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

row_idx = 2
for key, display in GROUPS:
    acc = accent(key)
    # Group separator row
    ws1.row_dimensions[row_idx].height = 18
    sep = ws1.cell(row_idx, 1, display.upper())
    sep.font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    sep.fill  = fill(acc)
    sep.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
    row_idx += 1

    for step in range(10):
        token  = f"{key}/{step}"
        lhex   = ld[key][str(step)]["value"]
        dhex   = dd[key][str(step)]["value"]
        oname  = old_name(key, step)
        pinned = step == 5

        ws1.row_dimensions[row_idx].height = 22
        row_fill = step5_fill if pinned else PatternFill()

        def cell_write(col, val, *, bold=False, center=False, italic=False, bkgd=None):
            c = ws1.cell(row_idx, col, val)
            c.font = Font(name="Arial", size=10, bold=bold, italic=italic)
            c.alignment = Alignment(horizontal="center" if center else "left",
                                    vertical="center", indent=0 if center else 1)
            if bkgd:  c.fill = fill(bkgd)
            elif pinned: c.fill = step5_fill
            return c

        cell_write(1, display, italic=True)
        cell_write(2, token)
        cell_write(3, step, center=True)
        cell_write(4, oname)
        cell_write(5, lhex)
        # Light swatch
        ls = ws1.cell(row_idx, 6, "\u2605" if pinned else "")
        ls.fill = fill(lhex)
        ls.alignment = Alignment(horizontal="center", vertical="center")
        ls.font = Font(name="Arial", size=10, color="FFFFFFFF" if pinned else "FF" + lhex.lstrip('#'))
        cell_write(7, dhex)
        # Dark swatch
        ds = ws1.cell(row_idx, 8, "\u2605" if pinned else "")
        ds.fill = fill(dhex)
        ds.alignment = Alignment(horizontal="center", vertical="center")
        ds.font = Font(name="Arial", size=10, color="FFFFFFFF" if pinned else "FF" + dhex.lstrip('#'))
        # Pinned col
        if pinned:
            pc = ws1.cell(row_idx, 9, "\u2605 original")
            pc.fill = pin_fill
            pc.font = Font(name="Arial", size=10, bold=True)
            pc.alignment = Alignment(horizontal="center", vertical="center")
        else:
            ws1.cell(row_idx, 9, "")
            if pinned: ws1.cell(row_idx, 9).fill = step5_fill

        row_idx += 1

# ─── Sheet 2: Active Colors ────────────────────
ws2 = wb.create_sheet("Active Colors")
ws2.freeze_panes = "A3"

ws2.append(["Colors actively in use \u2014 Old Name \u2192 New Mantine Token"])
ws2.cell(1,1).font = Font(name="Arial", bold=True, size=12)
ws2.merge_cells("A1:H1")
ws2.row_dimensions[1].height = 24

ws2.append(["Only colors that existed in the old design system. New interpolated steps excluded."])
ws2.cell(2,1).font = Font(name="Arial", italic=True, size=9)
ws2.merge_cells("A2:H2")
ws2.row_dimensions[2].height = 18

ws2.append(["Group","Old Name","New Token","Step","Light Hex","Light Swatch","Dark Hex","Dark Swatch"])
for col, w in enumerate([14,24,14,6,12,10,12,10], start=1):
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.row_dimensions[3].height = 22
for c in range(1,9):
    cell = ws2.cell(3, c)
    cell.font  = hdr_white
    cell.fill  = hdr_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

active_rows = [(k,d,s) for k,d in GROUPS for s in range(10)
               if old_name(k,s) not in ("\u2014","NEW")]

ridx = 4
last_key = None
for key, display, step in active_rows:
    if key != last_key:
        acc = accent(key)
        ws2.row_dimensions[ridx].height = 18
        sep = ws2.cell(ridx, 1, display.upper())
        sep.font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        sep.fill  = fill(acc)
        sep.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws2.merge_cells(start_row=ridx,start_column=1,end_row=ridx,end_column=8)
        ridx += 1
        last_key = key

    lhex = ld[key][str(step)]["value"]
    dhex = dd[key][str(step)]["value"]
    oname = old_name(key, step)
    token = f"{key}/{step}"
    pinned = step == 5

    ws2.row_dimensions[ridx].height = 22
    for c,v in [(1,display),(2,oname),(3,token),(4,step),(5,lhex),(7,dhex)]:
        cell = ws2.cell(ridx, c, v)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center" if c==4 else "left", vertical="center", indent=1)
        if pinned: cell.fill = step5_fill

    for col_idx, hex_val in [(6,lhex),(8,dhex)]:
        sc = ws2.cell(ridx, col_idx, "\u2605" if pinned else "")
        sc.fill = fill(hex_val)
        sc.alignment = Alignment(horizontal="center", vertical="center")
        sc.font = Font(name="Arial", size=10, color="FFFFFFFF" if pinned else "FF" + hex_val.lstrip('#'))

    ridx += 1

# ─── Sheet 3: Mantine Config ──────────────────
ws3 = wb.create_sheet("Mantine Config")
ws3.column_dimensions["A"].width = 110

note = ("Paste into createTheme({ colors: { ... } }) in your Mantine theme file. "
        "Index 5 = original 500 value (\u2605). 10 values per color, index 0\u20139.")
ws3.append([note])
ws3.cell(1,1).font = Font(name="Arial", italic=True, size=9)
ws3.row_dimensions[1].height = 18
ws3.append([])
ws3.row_dimensions[2].height = 6

ridx = 3
for key, display in GROUPS:
    vals = ", ".join(f"'{ld[key][str(i)]['value']}'" for i in range(10))
    line = f"  {key}: [{vals}],"
    ws3.append([f"// {display}"])
    ws3.cell(ridx,1).font = Font(name="Arial", bold=True, size=10, color="555577")
    ws3.row_dimensions[ridx].height = 18
    ridx += 1

    ws3.append([line])
    ws3.cell(ridx,1).font = Font(name="Courier New", size=10)
    ws3.row_dimensions[ridx].height = 18
    ridx += 1
    ws3.append([])
    ws3.row_dimensions[ridx].height = 6
    ridx += 1

wb.save(OUT/"color-system-mantine-expansion.xlsx")
print("✓ color-system-mantine-expansion.xlsx built")

# ─── Verification ─────────────────────────────
CHECKS = [
    ("blue","2","#82B8EB"), ("blue","5","#0471D7"),
    ("gray","1","#E3E3E3"), ("gray","2","#C7C7C7"), ("gray","5","#727272"),
    ("orange","2","#F2C7AF"), ("orange","5","#E48F5E"),
    ("purple","5","#A462EF"), ("green","5","#419D64"),
    ("yellow","5","#DBBB5A"), ("red","5","#E4675E"),
    ("magenta","5","#AC037A"),
]

print("\n=== VERIFICATION ===")
print(f"{'Token':<15} {'Expected':<12} {'Actual':<12} {'Pass?'}")
print("-"*46)
all_pass = True
for group, step, expected in CHECKS:
    actual = ld[group][step]["value"].upper()
    ok = actual == expected.upper()
    if not ok: all_pass = False
    print(f"{group+'/'+step:<15} {expected:<12} {actual:<12} {'✓' if ok else '✗'}")

print(f"\n{'All 12 ✓ — verification passed' if all_pass else '✗ FAILURES — check anchor values'}")
